"""
Excel'den resimleri DOĞRU eşleştirme ile çıkarır.
VML Drawing'deki satır bilgisini kullanarak resim-personel eşleşmesi yapar.
"""

import zipfile
import os
import re
import shutil
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from PIL import Image
import io
import psycopg2

EXCEL_FILE = "Kitap1.xlsx"
OUTPUT_DIR = "personnel_images_correct"
UPLOAD_DIR = "backend/uploads/avatars"

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "database": "eventflow",
    "user": "postgres",
    "password": "518518Erkan"
}

def clean_filename(name):
    """Dosya adı için geçersiz karakterleri temizle"""
    tr_map = {
        'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G',
        'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S',
        'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C'
    }
    for tr, en in tr_map.items():
        name = name.replace(tr, en)
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '')
    return name.strip()

def get_row_image_mapping():
    """VML'den satır-resim eşleşmesini çıkar"""
    row_image_map = {}
    
    with zipfile.ZipFile(EXCEL_FILE, 'r') as z:
        # Relations dosyasını oku
        rels_content = z.read('xl/drawings/_rels/vmlDrawing1.vml.rels').decode('utf-8')
        rels_root = ET.fromstring(rels_content)
        ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        
        rel_to_image = {}
        for rel in rels_root.findall('.//r:Relationship', ns):
            rid = rel.get('Id')
            target = rel.get('Target')
            if rid and target and 'image' in target.lower():
                image_name = os.path.basename(target)
                rel_to_image[rid] = image_name
        
        # VML'den shape'leri parse et
        vml_content = z.read('xl/drawings/vmlDrawing1.vml').decode('utf-8')
        shape_blocks = re.split(r'(?=<v:shape)', vml_content)
        
        for block in shape_blocks:
            if '<v:shape' not in block:
                continue
            
            row_match = re.search(r'<x:Row>(\d+)</x:Row>', block)
            relid_match = re.search(r'o:relid="(rId\d+)"', block)
            
            if row_match and relid_match:
                row = int(row_match.group(1))
                relid = relid_match.group(1)
                
                if relid in rel_to_image:
                    row_image_map[row] = rel_to_image[relid]
    
    return row_image_map

def get_excel_data():
    """Excel'den sicil ve isim bilgilerini oku"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    excel_data = {}
    for row_num in range(2, ws.max_row + 1):  # 1. satır başlık
        sicil = ws.cell(row=row_num, column=1).value
        isim = ws.cell(row=row_num, column=2).value
        
        if sicil and isim:
            sicil = str(sicil).strip()
            isim = str(isim).strip()
            # row_num - 1 çünkü VML 0-indexed
            excel_data[row_num - 1] = {"sicil": sicil, "isim": isim}
    
    return excel_data

def extract_and_save_images(row_image_map, excel_data):
    """Resimleri çıkar ve doğru isimle kaydet"""
    
    # Çıktı klasörlerini hazırla
    for dir_path in [OUTPUT_DIR, UPLOAD_DIR]:
        os.makedirs(dir_path, exist_ok=True)
    
    # Mevcut personnel_ dosyalarını temizle
    for f in os.listdir(UPLOAD_DIR):
        if f.startswith("personnel_"):
            os.remove(os.path.join(UPLOAD_DIR, f))
    
    saved = []
    
    with zipfile.ZipFile(EXCEL_FILE, 'r') as z:
        for row, image_file in row_image_map.items():
            if row not in excel_data:
                continue
            
            person = excel_data[row]
            sicil = person["sicil"]
            isim = person["isim"]
            
            # Resmi oku
            image_path = f"xl/media/{image_file}"
            try:
                image_data = z.read(image_path)
            except KeyError:
                print(f"  ✗ Resim bulunamadı: {image_path}")
                continue
            
            # PIL ile aç ve optimize et
            try:
                img = Image.open(io.BytesIO(image_data))
                
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')
                
                # Boyutu küçült
                max_size = (300, 300)
                img.thumbnail(max_size, Image.Resampling.LANCZOS)
                
                # Kaydet
                safe_name = clean_filename(isim)
                
                # personnel_images_correct klasörüne (referans için)
                ref_filename = f"{sicil}_{safe_name}.jpeg"
                ref_path = os.path.join(OUTPUT_DIR, ref_filename)
                img.save(ref_path, "JPEG", quality=85, optimize=True)
                
                # backend/uploads/avatars klasörüne
                upload_filename = f"personnel_{sicil}.jpeg"
                upload_path = os.path.join(UPLOAD_DIR, upload_filename)
                img.save(upload_path, "JPEG", quality=85, optimize=True)
                
                saved.append({
                    "sicil": sicil,
                    "isim": isim,
                    "avatar_path": f"/uploads/avatars/{upload_filename}"
                })
                
            except Exception as e:
                print(f"  ✗ {sicil} - {isim}: {e}")
    
    return saved

def update_database(saved_images):
    """Veritabanını güncelle"""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    # Önce tüm avatar'ları temizle
    cur.execute('UPDATE staff SET avatar = NULL')
    
    updated = 0
    for item in saved_images:
        cur.execute(
            'UPDATE staff SET avatar = %s WHERE "sicilNo" = %s',
            (item["avatar_path"], item["sicil"])
        )
        if cur.rowcount > 0:
            updated += 1
    
    conn.commit()
    cur.close()
    conn.close()
    
    return updated

def main():
    print("=" * 60)
    print("EXCEL'DEN RESİM ÇIKARMA (DOĞRU EŞLEŞTİRME)")
    print("=" * 60)
    
    # 1. VML'den satır-resim eşleşmesini al
    print("\n1. VML analiz ediliyor...")
    row_image_map = get_row_image_mapping()
    print(f"   {len(row_image_map)} resim-satır eşleşmesi bulundu")
    
    # 2. Excel'den personel bilgilerini al
    print("\n2. Excel okunuyor...")
    excel_data = get_excel_data()
    print(f"   {len(excel_data)} personel kaydı bulundu")
    
    # 3. Resimleri çıkar ve kaydet
    print("\n3. Resimler çıkarılıyor...")
    saved = extract_and_save_images(row_image_map, excel_data)
    print(f"   {len(saved)} resim kaydedildi")
    
    # 4. Veritabanını güncelle
    print("\n4. Veritabanı güncelleniyor...")
    updated = update_database(saved)
    print(f"   {updated} kayıt güncellendi")
    
    # Örnek eşleşmeleri göster
    print("\n" + "=" * 60)
    print("ÖRNEK EŞLEŞMELERİ (ilk 10):")
    for item in saved[:10]:
        print(f"  {item['sicil']} - {item['isim']}")
    
    print("\n" + "=" * 60)
    print("TAMAMLANDI!")
    print(f"  - Çıkarılan resim: {len(saved)}")
    print(f"  - Güncellenen DB kaydı: {updated}")
    print("=" * 60)

if __name__ == "__main__":
    main()
