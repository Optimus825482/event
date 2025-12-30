"""
Excel'den personel resimlerini DOĞRU eşleştirme ile çıkarıp veritabanına yükleyen script.
Openpyxl ile hücre koordinatlarını kullanarak resim-satır eşleşmesi yapılır.
"""

import os
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io
import psycopg2

# Konfigürasyon
EXCEL_FILE = "Kitap1.xlsx"
OUTPUT_DIR = "backend/uploads/avatars"
DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "database": "eventflow",
    "user": "postgres",
    "password": "518518Erkan"
}

# Sicil No sütunu (A sütunu = 1)
SICIL_COLUMN = 1
# Resim sütunu (genellikle A veya B sütununda gömülü)
IMAGE_COLUMN_RANGE = (1, 3)  # A-C sütunları arasında ara

def clean_output_dir():
    """Çıktı klasörünü temizle"""
    if os.path.exists(OUTPUT_DIR):
        # Sadece personnel_ ile başlayan dosyaları sil
        for f in os.listdir(OUTPUT_DIR):
            if f.startswith("personnel_"):
                os.remove(os.path.join(OUTPUT_DIR, f))
    else:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"✓ Çıktı klasörü hazırlandı: {OUTPUT_DIR}")

def get_image_row(image, ws):
    """Resmin hangi satırda olduğunu bul"""
    anchor = image.anchor
    if hasattr(anchor, '_from'):
        return anchor._from.row + 1  # 0-indexed to 1-indexed
    elif hasattr(anchor, 'row'):
        return anchor.row + 1
    return None

def extract_images_from_excel():
    """Excel'den resimleri çıkar ve sicil no ile eşleştir"""
    print(f"Excel dosyası yükleniyor: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Tüm resimleri ve konumlarını al
    images_by_row = {}
    
    if hasattr(ws, '_images'):
        for img in ws._images:
            row = get_image_row(img, ws)
            if row:
                images_by_row[row] = img
                print(f"  Resim bulundu: Satır {row}")
    
    print(f"Toplam {len(images_by_row)} resim bulundu")
    
    # Sicil numaralarını oku ve eşleştir
    results = []
    for row_num in range(2, ws.max_row + 1):  # 1. satır başlık
        sicil_cell = ws.cell(row=row_num, column=SICIL_COLUMN)
        sicil_no = str(sicil_cell.value).strip() if sicil_cell.value else None
        
        if not sicil_no or sicil_no == "None":
            continue
        
        # Bu satırda resim var mı?
        if row_num in images_by_row:
            img = images_by_row[row_num]
            results.append({
                "row": row_num,
                "sicil_no": sicil_no,
                "image": img
            })
    
    print(f"Eşleşen {len(results)} resim-sicil çifti bulundu")
    return results, wb

def save_images(image_data_list):
    """Resimleri dosya sistemine kaydet"""
    saved = []
    
    for item in image_data_list:
        sicil_no = item["sicil_no"]
        img = item["image"]
        
        try:
            # Resim verisini al
            img_data = img._data()
            
            # PIL ile aç
            pil_img = Image.open(io.BytesIO(img_data))
            
            # Dosya adı
            ext = "png" if pil_img.format == "PNG" else "jpeg"
            filename = f"personnel_{sicil_no}.{ext}"
            filepath = os.path.join(OUTPUT_DIR, filename)
            
            # Kaydet (optimize et)
            if pil_img.mode in ('RGBA', 'P'):
                pil_img = pil_img.convert('RGB')
            
            # Boyutu küçült (max 300x300)
            max_size = (300, 300)
            pil_img.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            pil_img.save(filepath, "JPEG", quality=85, optimize=True)
            
            saved.append({
                "sicil_no": sicil_no,
                "filename": filename,
                "filepath": f"/uploads/avatars/{filename}"
            })
            print(f"  ✓ {sicil_no} -> {filename}")
            
        except Exception as e:
            print(f"  ✗ {sicil_no} kaydedilemedi: {e}")
    
    return saved

def update_database(saved_images):
    """Veritabanındaki avatar alanlarını güncelle"""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    
    updated = 0
    not_found = []
    
    for item in saved_images:
        sicil_no = item["sicil_no"]
        avatar_path = item["filepath"]
        
        cur.execute(
            'UPDATE staff SET avatar = %s WHERE "sicilNo" = %s',
            (avatar_path, sicil_no)
        )
        
        if cur.rowcount > 0:
            updated += 1
        else:
            not_found.append(sicil_no)
    
    conn.commit()
    cur.close()
    conn.close()
    
    print(f"\n✓ Veritabanı güncellendi: {updated} kayıt")
    if not_found:
        print(f"⚠ Bulunamayan sicil numaraları: {not_found[:10]}...")
    
    return updated

def main():
    print("=" * 60)
    print("PERSONEL AVATAR ÇIKARMA VE YÜKLEME")
    print("=" * 60)
    
    # 1. Çıktı klasörünü temizle
    clean_output_dir()
    
    # 2. Excel'den resimleri çıkar
    image_data, wb = extract_images_from_excel()
    
    if not image_data:
        print("\n⚠ Excel'de gömülü resim bulunamadı!")
        print("Alternatif yöntem deneniyor: Satır bazlı resim arama...")
        
        # Alternatif: Her satır için manuel kontrol
        # Bu durumda resimlerin ayrı dosyalarda olması gerekir
        return
    
    # 3. Resimleri kaydet
    print("\nResimler kaydediliyor...")
    saved = save_images(image_data)
    
    # 4. Veritabanını güncelle
    print("\nVeritabanı güncelleniyor...")
    updated = update_database(saved)
    
    print("\n" + "=" * 60)
    print(f"TAMAMLANDI!")
    print(f"  - Çıkarılan resim: {len(saved)}")
    print(f"  - Güncellenen kayıt: {updated}")
    print("=" * 60)

if __name__ == "__main__":
    main()
